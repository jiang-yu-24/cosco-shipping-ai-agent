"""
工具函数定义模块 — 业务逻辑占位层
=====================================
此文件为业务逻辑占位，后续可替换为：
  - 合同/提单智能解析（OCR + NLP）
  - 散货船期/运价数据中台 API 调用
  - 港口拥堵指数、AIS 船舶轨迹查询
  - 企业内部 RAG 知识库检索
当前版本仅提供两个演示工具用于跑通 Agent 的 ReAct 闭环。
"""

import csv
import io
import json
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

# ============================================================
# 时区配置
# ============================================================
# 北京时间（东八区）— Streamlit Cloud 服务器使用 UTC 时区，
# 因此必须显式指定时区，否则 datetime.now() 会慢 8 小时
_CST = timezone(timedelta(hours=8), name="Asia/Shanghai")

# ============================================================
# 上传文件内容暂存区
# ============================================================
# app.py 在上传文件后将解析后的文本存储于此，
# Agent 工具 search_file_content 可检索该内容
_uploaded_file_content: str = ""
_uploaded_file_name: str = ""


def set_uploaded_file(content: str, filename: str) -> None:
    """供 app.py 调用，将已解析的文件内容存入模块全局变量。"""
    global _uploaded_file_content, _uploaded_file_name
    _uploaded_file_content = content
    _uploaded_file_name = filename


def get_uploaded_file_info() -> tuple:
    """返回 (content, filename) 供 app.py 读取。"""
    return _uploaded_file_content, _uploaded_file_name


# ============================================================
# 文件解析函数
# ============================================================

def parse_file_content(file_bytes: bytes, filename: str, max_chars: int = 8000) -> str:
    """
    根据文件扩展名选择解析器，提取文本内容。

    支持格式：PDF、Excel(.xlsx)、CSV、TXT
    超过 max_chars 时截断并追加提示。
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    try:
        if ext == "pdf":
            text = _parse_pdf(file_bytes)
        elif ext in ("xlsx", "xls"):
            text = _parse_excel(file_bytes)
        elif ext == "csv":
            text = _parse_csv(file_bytes)
        elif ext == "txt":
            text = file_bytes.decode("utf-8", errors="replace")
        else:
            return f"❌ 暂不支持 .{ext} 文件格式，请上传 PDF、Excel、CSV 或 TXT 文件。"

        if not text.strip():
            return "⚠️ 文件中未提取到可读文本内容。"

        if len(text) > max_chars:
            text = text[:max_chars] + f"\n\n... (文件内容已截断，共 {len(text)} 字符，仅展示前 {max_chars} 字符)"

        return text

    except Exception as e:
        return f"❌ 文件解析失败：{str(e)}"


def _parse_pdf(file_bytes: bytes) -> str:
    """从 PDF 二进制数据中提取文本。"""
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(file_bytes))
    parts = []
    for i, page in enumerate(reader.pages):
        page_text = page.extract_text()
        if page_text:
            parts.append(f"--- 第{i+1}页 ---\n{page_text}")
    return "\n\n".join(parts)


def _parse_excel(file_bytes: bytes) -> str:
    """从 Excel 二进制数据中提取所有工作表的文本。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"--- 工作表: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(v) if v is not None else "" for v in row]
            if any(v for v in row_vals):
                parts.append(" | ".join(row_vals))
    return "\n".join(parts)


def _parse_csv(file_bytes: bytes) -> str:
    """从 CSV 二进制数据中提取文本。"""
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    parts = []
    for row in reader:
        parts.append(" | ".join(row))
    return "\n".join(parts)


# ============================================================
# 工具函数定义区
# ============================================================

def get_current_time() -> str:
    """
    获取当前北京时间（东八区）。
    返回格式化的日期时间字符串，用于演示 Agent 调用本地工具的能力。
    """
    now = datetime.now(_CST)
    # 使用中文友好的日期时间格式
    return now.strftime("%Y年%m月%d日 %H:%M:%S (星期%w) 北京时间")


def query_shipping_schedule(route: str) -> str:
    """
    从100+条散货船期数据集中检索匹配的船期信息。

    支持按航线、港口、货种、船名、区域等关键词灵活检索，
    不再依赖硬编码数据。检索逻辑：关键词命中越多，排名越靠前。

    参数:
        route: str - 查询描述，如"西澳-青岛"、"天津铁矿石"、"巴西大豆"、"致远号"
    """
    import json
    import os

    # 加载船期数据集
    data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "data", "shipping_schedules.json")
    try:
        with open(data_path, "r", encoding="utf-8") as f:
            schedules = json.load(f)
    except Exception:
        return "⚠️ 船期数据加载失败，请联系管理员。"

    if not route.strip():
        return "⚠️ 请输入查询条件，如航线、港口、货种或船名。"

    # 关键词分词（按空格和常见分隔符拆分）
    keywords = route.strip().lower().replace("、", " ").replace("，", " ").replace(",", " ").split()
    if not keywords:
        keywords = [route.strip().lower()]

    # 对每条船期计算匹配分
    scored = []
    for s in schedules:
        text = f"{s['route']} {s['port_from']} {s['port_to']} {s['region']} {s['cargo']} {s['vessel']} {s['status']}".lower()
        score = 0
        for kw in keywords:
            if kw in text:
                score += 1
            # 部分匹配加分
            if len(kw) >= 2:
                for i in range(len(kw) - 1):
                    if kw[i:i+2] in text:
                        score += 0.25
        if score > 0:
            scored.append((score, s))

    if not scored:
        # 无匹配时返回提示
        return (
            f"⚠️ 未找到与「{route}」匹配的船期数据。\n"
            f"当前数据库收录 {len(schedules)} 条船期，覆盖以下区域：\n"
            f"澳大利亚-中国、南美-中国、东南亚-中国、北美-中国、"
            f"西非-中国、南非-中国、远东-中国、印度-中国\n"
            f"支持的货种：铁矿石、煤炭/焦煤/动力煤、大豆/玉米/小麦、铝土矿、锰矿等\n"
            f"请尝试使用港口名、城市名、货种名或区域名重新查询。"
        )

    # 按匹配分降序，取前5条
    scored.sort(key=lambda x: x[0], reverse=True)
    top = scored[:5]

    if len(top) == 1:
        s = top[0][1]
        return (
            f"📍 航线：{s['route']}\n"
            f"🚢 船名：{s['vessel']}\n"
            f"📦 货种：{s['cargo']}\n"
            f"⚓ 装货港：{s['port_from']}\n"
            f"🏁 卸货港：{s['port_to']}\n"
            f"📅 预计离港：{s['departure']}\n"
            f"📅 预计到港：{s['arrival']}\n"
            f"⏱️ 航程：{s['duration']}\n"
            f"📡 状态：{s['status']}"
            + (f"\n📝 备注：{s['remark']}" if s['remark'] else "")
        )

    # 多条匹配
    result = f"🔍 「{route}」匹配到 {len(top)} 条船期（共 {len(scored)} 条相关）：\n\n"
    for i, (score, s) in enumerate(top, 1):
        result += (
            f"【{i}】{s['route']} | {s['vessel']} | {s['cargo']}\n"
            f"    {s['port_from']} → {s['port_to']} | {s['departure']} → {s['arrival']} | {s['status']}\n\n"
        )
    if len(scored) > 5:
        result += f"（还有 {len(scored) - 5} 条匹配未展示，请缩小查询范围）"
    else:
        result += "如需查看某条详情，请提供更具体的查询条件。"
    return result


def search_file_content(keyword: str) -> str:
    """
    在用户上传的文件中搜索关键词，返回匹配的段落。

    适用于用户上传了提单、合同、船期表等文件后，
    需要从文件中提取特定信息（如托运人、货量、港口等）。

    参数:
        keyword: str - 要搜索的关键词或短语
    """
    content, filename = get_uploaded_file_info()

    if not content:
        return "⚠️ 当前没有已上传的文件。请先在左侧上传一份文件（PDF/Excel/CSV/TXT）。"

    if not keyword.strip():
        return "⚠️ 请输入要搜索的关键词。"

    # 将内容按行拆分，搜索包含关键词的行及上下文
    lines = content.split("\n")
    matched_blocks = []
    kw_lower = keyword.lower()

    for i, line in enumerate(lines):
        if kw_lower in line.lower():
            # 取匹配行及上下各1行作为上下文
            start = max(0, i - 1)
            end = min(len(lines), i + 2)
            block = "\n".join(lines[start:end])
            matched_blocks.append(block)

    if not matched_blocks:
        return (
            f"📄 在文件「{filename}」中未找到与「{keyword}」相关的内容。\n"
            f"建议尝试其他关键词，或直接询问我关于文件内容的概括性问题。"
        )

    # 最多返回前5个匹配块
    result = f"📄 在文件「{filename}」中找到 {len(matched_blocks)} 处「{keyword}」相关匹配：\n\n"
    for j, block in enumerate(matched_blocks[:5], 1):
        result += f"【匹配 {j}】\n{block}\n\n"

    if len(matched_blocks) > 5:
        result += f"... (还有 {len(matched_blocks) - 5} 处匹配未展示，建议缩小搜索范围)"

    return result


def generate_document(doc_type: str, title: str = "", content: str = "",
                      route: str = "", vessel: str = "", departure: str = "",
                      arrival: str = "", cargo: str = "",
                      consignor: str = "", consignee: str = "",
                      recipient: str = "") -> str:
    """
    生成 PDF 文档（船期确认函/航运报告/通用公文/通用格式/项目方案）。

    当用户要求生成文档时调用此工具。红头公文类用 schedule/report/official，
    非正式文档用 generic，数字化项目方案用 proposal。生成后可下载 PDF 文件。

    参数:
        doc_type: 文档类型 — "schedule"（船期确认函）、"report"（航运报告）、"official"（通用公文）、"generic"（通用格式）、"proposal"（数字化项目方案）
        title: 文档标题
        content: 正文内容（换行分段）
        route: [schedule] 航线
        vessel: [schedule] 船名
        departure: [schedule] 离港时间
        arrival: [schedule] 到港时间
        cargo: [schedule] 货种货量
        consignor: [schedule] 托运人
        consignee: [schedule] 收货人
        recipient: [official] 主送单位
    """
    from pdf_utils import (
        generate_schedule_confirmation,
        generate_shipping_report,
        generate_official_document,
        generate_generic_pdf,
        generate_proposal,
        store_pdf,
    )

    now = datetime.now(_CST)
    ts = now.strftime("%Y%m%d_%H%M%S")

    try:
        if doc_type == "schedule":
            pdf_bytes = generate_schedule_confirmation(
                route=route or "待指定",
                vessel=vessel or "待指定",
                departure=departure or "待确认",
                arrival=arrival or "待确认",
                cargo=cargo or "待指定",
                consignor=consignor or "待填写",
                consignee=consignee or "待填写",
            )
            filename = f"船期确认函_{route or '通用'}_{ts}.pdf"

        elif doc_type == "report":
            pdf_bytes = generate_shipping_report(
                title=title or "航运分析报告",
                content=content or "（无正文内容）",
            )
            filename = f"航运报告_{ts}.pdf"

        elif doc_type == "official":
            pdf_bytes = generate_official_document(
                title=title or "通知",
                content=content or "（无正文内容）",
                recipient=recipient or "",
            )
            filename = f"公文_{title or '通知'}_{ts}.pdf"

        elif doc_type == "generic":
            pdf_bytes = generate_generic_pdf(
                title=title or "文档",
                content=content or "（无正文内容）",
            )
            filename = f"{title or '文档'}_{ts}.pdf"

        elif doc_type == "proposal":
            pdf_bytes = generate_proposal(
                title=title or "数字化项目方案",
                project_name=route or "",    # 复用 route 参数传递项目名称
                department=consignor or "",  # 复用 consignor 参数传递申报单位
                content=content or "（无正文内容）",
            )
            filename = f"项目方案_{title or '未命名'}_{ts}.pdf"

        else:
            return (
                f"❌ 不支持的文档类型「{doc_type}」。"
                f"可选类型：schedule（船期确认函）、report（航运报告）、"
                f"official（通用公文）、generic（通用格式）、proposal（项目方案）"
            )

        store_pdf(pdf_bytes, filename)
        return (
            f"✅ PDF 文档已生成：{filename}（{len(pdf_bytes) / 1024:.1f} KB）\n"
            f"请告知用户点击页面上的「下载 PDF」按钮获取文件。"
        )

    except Exception as e:
        return f"❌ PDF 生成失败：{str(e)}"


def compose_email(subject: str, body: str, recipient: str = "") -> str:
    """
    编写邮件，生成可复制主题和正文的文本框。
    支持依次调用多次生成多封邮件。

    参数:
        subject: 邮件主题
        body: 邮件正文（可含换行符）
        recipient: 收件人（可选）
    """
    global _email_queue
    _email_queue.append((subject, body, recipient))
    return f"邮件「{subject}」已编写，可在结果区直接复制。"


_email_queue: list = []  # [(subject, body, recipient), ...]


def get_emails() -> list:
    """获取所有待展示的邮件列表，读取后清空。"""
    global _email_queue
    result = _email_queue[:]
    _email_queue = []
    return result


# ============================================================
# 工具注册表 — 供 agent_core.py 和 app.py 使用
# ============================================================

# TOOL_DESCRIPTIONS: 符合 OpenAI Function Calling 规范的工具定义列表
# 每个工具包含 name（函数名）、description（功能说明）、parameters（参数JSON Schema）
TOOL_DESCRIPTIONS: List[Dict[str, Any]] = [
    {
        "type": "function",
        "function": {
            "name": "get_current_time",
            "description": "获取当前北京时间（东八区）。当用户询问'现在几点'、'今天几号'、'当前时间'时调用此工具。",
            "parameters": {
                "type": "object",
                "properties": {},
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "query_shipping_schedule",
            "description": (
                "查询散货船期数据库（100+条记录）。"
                "支持按航线、港口名、城市名、货种、船名、区域等任意关键词检索。"
                "覆盖澳大利亚、南美、东南亚、北美、西非、南非、远东至中国航线。"
                "货种含铁矿石、煤炭、大豆、铝土矿、锰矿等。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "route": {
                        "type": "string",
                        "description": "查询关键词，如'西澳-青岛'、'天津铁矿石'、'巴西大豆'、'致远号'、'动力煤'等",
                    },
                },
                "required": ["route"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_file_content",
            "description": (
                "读取并分析用户已上传的文件内容。支持 PDF、Excel、CSV、TXT 格式。"
                "当用户询问文件中具体信息（如某字段值、某关键词出现位置、数据汇总等）时调用。"
                "可指定关键词精确定位，也可不指定关键词返回全文概览。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "keyword": {
                        "type": "string",
                        "description": "要搜索的关键词或短语，例如'托运人'、'装货港'、'天津'",
                    },
                },
                "required": ["keyword"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "generate_document",
            "description": (
                "生成 PDF 文档。根据用户意图选择类型：\n"
                "- 用户提到「船期」「确认函」「船运安排」→ 用 schedule\n"
                "- 用户提到「分析报告」「航运报告」「周报」→ 用 report\n"
                "- 用户提到「通知」「函件」「请示」「批复」「决定」→ 用 official\n"
                "- 用户提到「说明书」「教程」「手册」「总结」「指南」→ 用 generic\n"
                "- 用户提到「项目方案」「建设方案」「实施方案」「立项」「规划」「可研」→ 用 proposal\n"
                "proposal 会生成央企标准格式：封面+分章节正文+签审页。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "doc_type": {
                        "type": "string",
                        "description": "schedule=船期确认函 report=航运报告 official=公文 generic=简洁排版 proposal=项目方案",
                        "enum": ["schedule", "report", "official", "generic", "proposal"],
                    },
                    "title": {
                        "type": "string",
                        "description": "文档标题",
                    },
                    "content": {
                        "type": "string",
                        "description": "正文内容，使用换行符分隔段落",
                    },
                    "route": {
                        "type": "string",
                        "description": "[schedule] 航线，如'西澳-青岛'",
                    },
                    "vessel": {
                        "type": "string",
                        "description": "[schedule] 承运船舶名称",
                    },
                    "departure": {
                        "type": "string",
                        "description": "[schedule] 预计离港时间",
                    },
                    "arrival": {
                        "type": "string",
                        "description": "[schedule] 预计到港时间",
                    },
                    "cargo": {
                        "type": "string",
                        "description": "[schedule] 货种及货量",
                    },
                    "consignor": {
                        "type": "string",
                        "description": "[schedule] 托运人名称",
                    },
                    "consignee": {
                        "type": "string",
                        "description": "[schedule] 收货人名称",
                    },
                    "recipient": {
                        "type": "string",
                        "description": "[official] 主送单位",
                    },
                },
                "required": ["doc_type"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "compose_email",
            "description": (
                "编写邮件。生成可点击复制的主题和正文。"
                "当用户要求「写邮件」「起草邮件」「帮我回复客户」等场景时调用。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "subject": {
                        "type": "string",
                        "description": "邮件主题（不含'主题：'前缀）",
                    },
                    "body": {
                        "type": "string",
                        "description": "邮件正文（不含'正文：'前缀，可含换行）",
                    },
                    "recipient": {
                        "type": "string",
                        "description": "收件人邮箱或名称（可选）",
                    },
                },
                "required": ["subject", "body"],
            },
        },
    },
]

# TOOL_MAPPING: 工具名 -> 实际Python函数的映射字典
# agent_core.py 通过此字典查找并执行本地函数
TOOL_MAPPING: Dict[str, Any] = {
    "get_current_time": get_current_time,
    "query_shipping_schedule": query_shipping_schedule,
    "search_file_content": search_file_content,
    "generate_document": generate_document,
    "compose_email": compose_email,
}

# TOOL_NAMES: 工具名称列表，供 app.py 侧边栏展示
TOOL_NAMES: List[str] = [t["function"]["name"] for t in TOOL_DESCRIPTIONS]

# TOOL_DISPLAY_NAMES: 工具函数名 -> 自然语言展示名
# app.py 侧边栏使用此映射展示服务能力，对用户隐藏内部函数名
TOOL_DISPLAY_NAMES: Dict[str, str] = {
    "search_file_content": "读取文件内容",
    "generate_document": "生成PDF文件",
    "compose_email": "编写邮件",
}
